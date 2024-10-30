import dotenv
import os
import openpyxl.workbook
import pymysql
import openpyxl
from pathlib import Path
import functions_

dotenv.load_dotenv()




# CONFIGURATION THE DATABASE.

connect = pymysql.connect(
    host = os.environ['SQL_HOST'],
    user= os.environ['SQL_USER'],
    password= os.environ['SQL_PASSWORD'],
    database= os.environ['SQL_DATABASE'],
)

cursor = connect.cursor()


# RELOAD TABLE.

cursor.execute(
    'DROP TABLE subjects'
)

cursor.execute(
    'DROP TABLE matrix'
)

# Creating the matrix table.
cursor.execute(
        'CREATE TABLE IF NOT EXISTS matrix '
        '('
        'id SMALLINT UNSIGNED PRIMARY KEY AUTO_INCREMENT, '
        'name_matter VARCHAR(250) NOT NULL '
        '); '               
               )

# Creating the subjects_table.
cursor.execute(
        'CREATE TABLE IF NOT EXISTS subjects '
        '('
        'id SMALLINT UNSIGNED PRIMARY KEY AUTO_INCREMENT, '
        'id_matrix SMALLINT UNSIGNED NOT NULL,'
        'name_subject VARCHAR(250) NOT NULL, '
        'conclusion_date DATE NULL, '
        'FOREIGN KEY (id_matrix) REFERENCES matrix(id) '
        'ON DELETE CASCADE ON UPDATE CASCADE'
        '); '
        )




connect.commit()

# cursor.close()
# connect.close()

print('SQL Process conclude.')


# PART OF XLSX

FILE = Path(__file__).parent / 'XLSX_FILES' / 'CICLO DE ESTUDOS - POLÍCIAS - \
GERAL.xlsx'


# Loading the file. 

file = openpyxl.load_workbook(filename=FILE)


# Selecting the sheet to consult.

sheet_consult = file['MATRIZ_CONTEÚDO']

# Sending the values (matters) to the database.
for row in sheet_consult.iter_cols(min_col= 2, max_col= 2, min_row= 2,
max_row= 12, values_only= True):
    for value in row:
        
        cursor.execute(
            f'INSERT INTO matrix (name_matter) VALUES (%s)', value
        )

connect.commit()

# SENDING THE VALUES OF SUBJECTS TO THE DATABASE.

# Sending the values (subject: língua portuguesa) to the database.
functions_.capture_subjects(sheet_consult, 18, 168, 'name_subject', 1, cursor)

# Sending the values (subject: informática) to the database.
functions_.capture_subjects(sheet_consult, 171, 183, 'name_subject', 2, cursor)

# Sending the values (subject: raciocínio lógico) to the database.
functions_.capture_subjects(sheet_consult, 186, 192, 'name_subject', 3, cursor)

# Sending the values (subject: adm. pública e ética no serviço
# público) to the database.
functions_.capture_subjects(sheet_consult, 195, 208, 'name_subject', 4, cursor)

# Sending the values (subject: legislação específica) to the database.
functions_.capture_subjects(sheet_consult, 211, 214, 'name_subject', 5, cursor)

# Sending the values (subject: direito administrativo) to the database.
functions_.capture_subjects(sheet_consult, 217, 267, 'name_subject', 6, cursor)

# Sending the values (subject: direito constitucional) to the database.
functions_.capture_subjects(sheet_consult, 270, 315, 'name_subject', 7, cursor)

# Sending the values (subject: direitos humanos) to the database.
functions_.capture_subjects(sheet_consult, 318, 322, 'name_subject', 8, cursor)

# Sending the values (subject: direito penal) to the database.
functions_.capture_subjects(sheet_consult, 325, 357, 'name_subject', 9, cursor)

# Sending the values (subject: direito processual penal) to the database.
functions_.capture_subjects(sheet_consult, 395, 430, 'name_subject', 10, cursor)

# Sending the values (subject: legislação extravagante) to the database.
functions_.capture_subjects(sheet_consult, 360, 392, 'name_subject', 11, cursor)

connect.commit()

# SENDING THE VALUES OF CONCLUSION DATE TO THE DATABASE.
for row in sheet_consult.iter_rows(18, 430, 2, 4, values_only=True):
    name, _, date = row
    cursor.execute(
        'UPDATE subjects SET conclusion_date = %s WHERE name_subject = %s',
        (date, name)
    )
        
connect.commit()

cursor.close()
connect.close()

print('-'*15, 'SCRIPT CONCLUDED', '-'*15)
